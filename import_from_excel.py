import json
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List

if sys.version_info >= (3, 11):
    import tomllib
else:
    import tomli as tomllib


class ExcelImporter:
    def __init__(self, config_path: str = "config.toml"):
        with open(config_path, 'rb') as f:
            self.config = tomllib.load(f)

        self.trans_config = self.config['translation_settings']
        self.excel_config = self.config['excel_export']

    def excel_to_json(self, excel_file: str, output_folder: str):
        if not os.path.exists(excel_file):
            print(f"Error: Excel file '{excel_file}' does not exist")
            return

        print(f"Loading Excel file: {excel_file}")
        wb = load_workbook(excel_file, data_only=True)

        output_path = Path(output_folder)
        all_json_files = list(output_path.rglob('*.json'))

        for sheet_name in wb.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")
            ws = wb[sheet_name]

            file_path_cell = ws.cell(row=2, column=1).value

            if file_path_cell:
                json_filename = file_path_cell + '.json'
                json_path = Path(output_folder) / json_filename
                print(f"  Using FilePath from column A: {file_path_cell}")
            else:
                json_filename = sheet_name.replace('_', os.sep) + '.json'
                json_path = Path(output_folder) / json_filename

                if not json_path.exists():
                    matching_file = None
                    for json_file in all_json_files:
                        relative_path = json_file.relative_to(output_path)
                        truncated_name = str(relative_path.with_suffix('')).replace(os.sep, '_')[:31]

                        if truncated_name == sheet_name:
                            matching_file = json_file
                            break

                    if matching_file:
                        json_path = matching_file
                        print(f"  Matched truncated sheet name to: {json_path.name}")
                    else:
                        print(f"  Warning: No matching JSON found for sheet '{sheet_name}', skipping...")
                        continue

            if not json_path.exists():
                print(f"  Warning: Original JSON not found at {json_path}, skipping...")
                continue

            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            updates = self.read_sheet_data(ws)
            updated_count = self.apply_updates(data, updates)

            output_path = Path(output_folder) / json_filename
            os.makedirs(output_path.parent, exist_ok=True)

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)

            print(f"  Updated {updated_count} entries")
            print(f"  Saved to: {output_path}")

        print("\n✓ Import complete!")

    def read_sheet_data(self, ws) -> Dict:
        updates = {'blocks': {}, 'choices': {}}

        headers = self.excel_config['columns']
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header in headers:
                col_map[header] = col_idx

        for row_idx in range(2, ws.max_row + 1):
            block_idx_cell = ws.cell(row=row_idx, column=col_map.get('blockIdx', 1)).value

            if not block_idx_cell:
                continue

            block_idx_str = str(block_idx_cell)

            if '-C' in block_idx_str:
                choice_key = block_idx_str
                updates['choices'][choice_key] = {
                    'enText': ws.cell(row=row_idx, column=col_map.get('enText', 5)).value or '',
                    'QC': ws.cell(row=row_idx, column=col_map.get('QC', 6)).value or ''
                }
            else:
                try:
                    block_idx = int(block_idx_str)
                    updates['blocks'][block_idx] = {
                        'enName': ws.cell(row=row_idx, column=col_map.get('enName', 3)).value or '',
                        'enText': ws.cell(row=row_idx, column=col_map.get('enText', 5)).value or '',
                        'QC': ws.cell(row=row_idx, column=col_map.get('QC', 6)).value or ''
                    }
                except ValueError:
                    continue

        return updates

    def apply_updates(self, data: Dict, updates: Dict) -> int:
        updated_count = 0

        for block in data.get('text', []):
            block_idx = block.get('blockIdx')

            if block_idx in updates['blocks']:
                update_data = updates['blocks'][block_idx]

                qc_value = update_data.get('QC', '').strip()
                if qc_value:
                    if block.get('enText', '') != qc_value:
                        block['enText'] = qc_value
                        updated_count += 1
                else:
                    en_text = update_data.get('enText', '').strip()
                    if en_text and block.get('enText', '') != en_text:
                        block['enText'] = en_text
                        updated_count += 1

                en_name = update_data.get('enName', '').strip()
                if en_name and block.get('enName', '') != en_name:
                    block['enName'] = en_name
                    updated_count += 1

            if 'choices' in block and block['choices']:
                for choice_idx, choice in enumerate(block['choices']):
                    choice_key = f"{block_idx}-C{choice_idx + 1}"

                    if choice_key in updates['choices']:
                        choice_data = updates['choices'][choice_key]

                        qc_value = choice_data.get('QC', '').strip()
                        if qc_value:
                            if choice.get('enText', '') != qc_value:
                                choice['enText'] = qc_value
                                updated_count += 1
                        else:
                            en_text = choice_data.get('enText', '').strip()
                            if en_text and choice.get('enText', '') != en_text:
                                choice['enText'] = en_text
                                updated_count += 1

        return updated_count

    def import_qc_updates(self, excel_file: str = None):
        if excel_file is None:
            excel_file = self.excel_config['output_file']

        output_folder = self.trans_config['output_folder']
        self.excel_to_json(excel_file, output_folder)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Import QC updates from Excel to JSON')
    parser.add_argument('--excel', '-e', help='Excel file to import')
    parser.add_argument('--output', '-o', help='Output folder for updated JSON files')
    args = parser.parse_args()

    importer = ExcelImporter()

    if args.excel and args.output:
        importer.excel_to_json(args.excel, args.output)
    else:
        importer.import_qc_updates(args.excel)


if __name__ == "__main__":
    main()
