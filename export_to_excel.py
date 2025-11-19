import json
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

if sys.version_info >= (3, 11):
    import tomllib
else:
    import tomli as tomllib


class ExcelExporter:
    def __init__(self, config_path: str = "config.toml"):
        with open(config_path, 'rb') as f:
            self.config = tomllib.load(f)

        self.trans_config = self.config['translation_settings']
        self.excel_config = self.config['excel_export']

    def json_to_excel(self, input_folder: str, output_file: str):
        input_path = Path(input_folder)

        if not input_path.exists():
            print(f"Error: Input folder '{input_folder}' does not exist")
            return

        json_files = list(input_path.rglob('*.json'))

        if not json_files:
            print(f"No JSON files found in '{input_folder}'")
            return

        print(f"Found {len(json_files)} JSON files to export")

        wb = Workbook()
        wb.remove(wb.active)

        for json_file in json_files:
            print(f"Processing: {json_file}")
            self.add_json_to_workbook(wb, json_file, input_path)

        wb.save(output_file)
        print(f"\n✓ Excel file saved to: {output_file}")

    def add_json_to_workbook(self, wb: Workbook, json_path: Path, base_path: Path):
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        relative_path = json_path.relative_to(base_path)
        full_path_str = str(relative_path.with_suffix(''))
        sheet_name = full_path_str.replace(os.sep, '_')[:31]

        ws = wb.create_sheet(title=sheet_name)

        headers = ['FilePath'] + self.excel_config['columns']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        for block in data.get('text', []):
            ws.cell(row=row_idx, column=1, value=full_path_str)
            ws.cell(row=row_idx, column=2, value=block.get('blockIdx', ''))
            ws.cell(row=row_idx, column=3, value=block.get('jpName', ''))
            ws.cell(row=row_idx, column=4, value=block.get('enName', ''))
            ws.cell(row=row_idx, column=5, value=block.get('jpText', ''))
            ws.cell(row=row_idx, column=6, value=block.get('enText', ''))
            ws.cell(row=row_idx, column=7, value='')

            ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")

            row_idx += 1

            if 'choices' in block and block['choices']:
                for choice_idx, choice in enumerate(block['choices']):
                    ws.cell(row=row_idx, column=1, value=full_path_str)
                    ws.cell(row=row_idx, column=2, value=f"{block.get('blockIdx', '')}-C{choice_idx + 1}")
                    ws.cell(row=row_idx, column=3, value="[Choice]")
                    ws.cell(row=row_idx, column=4, value="[Choice]")
                    ws.cell(row=row_idx, column=5, value=choice.get('jpText', ''))
                    ws.cell(row=row_idx, column=6, value=choice.get('enText', ''))
                    ws.cell(row=row_idx, column=7, value='')

                    ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row=row_idx, column=6).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")

                    row_idx += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 50

        ws.freeze_panes = 'A2'

        print(f"  Added {row_idx - 2} rows to sheet '{sheet_name}'")

    def export_translated_files(self):
        output_folder = self.trans_config['output_folder']
        excel_file = self.excel_config['output_file']

        self.json_to_excel(output_folder, excel_file)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Export JSON translations to Excel')
    parser.add_argument('--input', '-i', help='Input folder containing JSON files')
    parser.add_argument('--output', '-o', help='Output Excel file path')
    args = parser.parse_args()

    exporter = ExcelExporter()

    if args.input and args.output:
        exporter.json_to_excel(args.input, args.output)
    else:
        exporter.export_translated_files()


if __name__ == "__main__":
    main()
